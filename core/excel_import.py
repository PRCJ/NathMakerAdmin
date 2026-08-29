import csv
import io
import os
import re
from typing import Any, Optional
from urllib.parse import unquote, urlparse

HEADER_ALIASES = {
    "productname": "productName",
    "product_name": "productName",
    "name": "productName",
    "product": "productName",
    "title": "productName",
    "description": "description",
    "desc": "description",
    "details": "description",
    "price": "price",
    "mrp": "price",
    "amount": "price",
    "rs": "price",
    "material": "material",
    "weight": "weight",
    "catalogue": "catalogue",
    "catalog": "catalogue",
    "collection": "catalogue",
    "catalogueid": "catalogue",
    "catalogue_id": "catalogue",
    "image": "image",
    "images": "image",
    "imageurl": "image",
    "imageurls": "image",
    "photo": "image",
    "photos": "image",
    "img": "image",
    "picture": "image",
    "filename": "image",
    "file": "image",
    "available": "available",
    "isavailable": "available",
    "show": "available",
}


def _norm_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_price(value: Any) -> float:
    raw = _cell_str(value)
    if not raw:
        return 0.0
    cleaned = re.sub(r"[₹,\s]", "", raw)
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def _parse_available(value: Any) -> bool:
    raw = _cell_str(value).lower()
    if raw in ("1", "true", "yes", "y", "available"):
        return True
    return False


def split_image_refs(value: Any) -> list[str]:
    raw = _cell_str(value)
    if not raw:
        return []
    parts = re.split(r"[\n,;|]+", raw)
    return [part.strip() for part in parts if part.strip()]


def is_image_url(ref: str) -> bool:
    return ref.lower().startswith("http://") or ref.lower().startswith("https://")


def name_from_image_ref(ref: str) -> str:
    path = unquote(urlparse(ref).path if is_image_url(ref) else ref)
    base = os.path.basename(path.rstrip("/")) or ref
    name = re.sub(r"\.[a-zA-Z0-9]+$", "", base)
    name = re.sub(r"[_-]+", " ", name).strip()
    return name or "New piece"


def _map_headers(headers: list[Any]) -> dict[int, str]:
    mapped = {}
    for idx, header in enumerate(headers):
        key = HEADER_ALIASES.get(_norm_header(header))
        if key:
            mapped[idx] = key
    return mapped


def _row_to_product(values: list[Any], header_map: dict[int, str], row_number: int) -> Optional[dict]:
    data = {
        "productName": "",
        "description": "",
        "price": 0.0,
        "material": "",
        "weight": "",
        "catalogue": "",
        "image": "",
        "available": False,
    }
    for idx, field in header_map.items():
        if idx >= len(values):
            continue
        data[field] = values[idx]

    image_refs = split_image_refs(data["image"])
    name = _cell_str(data["productName"])
    description = _cell_str(data["description"])
    material = _cell_str(data["material"])
    weight = _cell_str(data["weight"])
    catalogue = _cell_str(data["catalogue"])
    price = _parse_price(data["price"])

    if not name and image_refs:
        name = name_from_image_ref(image_refs[0])

    if not name and not image_refs and not description and price == 0 and not material and not weight:
        return None

    if not name:
        name = "New piece"

    return {
        "row": row_number,
        "productName": name,
        "description": description,
        "price": price,
        "material": material,
        "weight": weight,
        "catalogue": catalogue,
        "imageRefs": image_refs,
        "isAvailable": _parse_available(data["available"]),
    }


def parse_csv_bytes(content: bytes) -> list[dict]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return []
    header_map = _map_headers(rows[0])
    if not header_map:
        raise ValueError("No recognised columns. Use productName, price, description, material, weight, catalogue, image, available.")
    products = []
    for offset, row in enumerate(rows[1:], start=2):
        parsed = _row_to_product(row, header_map, offset)
        if parsed:
            products.append(parsed)
    return products


def parse_xlsx_bytes(content: bytes) -> list[dict]:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()
    if not rows:
        return []
    header_map = _map_headers(list(rows[0]))
    if not header_map:
        raise ValueError("No recognised columns. Use productName, price, description, material, weight, catalogue, image, available.")
    products = []
    for offset, row in enumerate(rows[1:], start=2):
        parsed = _row_to_product(list(row or []), header_map, offset)
        if parsed:
            products.append(parsed)
    return products


def parse_spreadsheet(filename: str, content: bytes) -> list[dict]:
    lower = (filename or "").lower()
    if lower.endswith(".csv"):
        return parse_csv_bytes(content)
    if lower.endswith(".xlsx") or lower.endswith(".xlsm"):
        return parse_xlsx_bytes(content)
    raise ValueError("Use an .xlsx or .csv file.")


def build_template_xlsx() -> bytes:
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "products"
    sheet.append(
        ["productName", "description", "price", "material", "weight", "catalogue", "image", "available"]
    )
    sheet.append(
        [
            "Gold Nath",
            "Handcrafted bridal nath",
            12500,
            "22K Gold",
            "4.2g",
            "Bridal",
            "https://example.com/nath.jpg",
            "no",
        ]
    )
    sheet.append(["", "", "", "", "", "", "ring-front.jpg", "no"])
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()
